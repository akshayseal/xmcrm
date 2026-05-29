import csv
import hashlib
import io
import json
import os
import re
import secrets
import sqlite3
from datetime import datetime, timedelta
from http import cookies
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook


ROOT = Path(__file__).parent
DB_PATH = Path(os.environ.get("DB_PATH", ROOT / "schoolcrm.sqlite3"))
SAMPLE_XLSX = Path("/Users/akshayseal/Downloads/Chandigarh Schools - Updated.xlsx")
FIELD_LIMIT = 30
ADMIN_SECRET = os.environ.get("ADMIN_SECRET", "XM-ADMIN-2026")

FIELDS = [
    "SCHOOL CODE",
    "SCHOOL NAME",
    "ADDRESS",
    "ADDRESS-2",
    "CITY",
    "DISTRICT",
    "PIN",
    "STATE/ UT",
    "BOARD",
    "BOARD AFFILIATION NO.",
    "UDISE NO",
    "GOVT/PRIVATE",
    "PRIORITY",
    "STRENGTH",
    "GRADE LEVEL UPTO WHICH STANDARD",
    "SCHOOL PHONE NO.",
    "SCHOOL MOBILE NO",
    "SCHOOL EMAIL ID",
    "WEBSITE",
    "PRINCIPAL NAME",
    "PRINCIPAL MOBILE NO.",
    "PRINCIPAL EMAIL ID",
    "SPoC/COORDINATOR NAME",
    "SPoC MOBILE NO",
    "SPoC EMAIL ID",
    "EXTRA PHONE NUMBER",
    "EXTRA EMAIL ID",
    "OLD PHONE",
    "OLD EMAIL ID",
    "PARTICIPATION",
]

STATE_CODES = {
    "Andaman and Nicobar Islands": "AN",
    "Andhra Pradesh": "AP",
    "Arunachal Pradesh": "AR",
    "Assam": "AS",
    "Bihar": "BR",
    "Chandigarh": "CH",
    "Chhattisgarh": "CG",
    "Delhi": "DL",
    "Goa": "GA",
    "Gujarat": "GJ",
    "Haryana": "HR",
    "Himachal Pradesh": "HP",
    "Jammu and Kashmir": "JK",
    "Jharkhand": "JH",
    "Karnataka": "KA",
    "Kerala": "KL",
    "Ladakh": "LA",
    "Lakshadweep": "LD",
    "Madhya Pradesh": "MP",
    "Maharashtra": "MH",
    "Manipur": "MN",
    "Meghalaya": "ML",
    "Mizoram": "MZ",
    "Nagaland": "NL",
    "Odisha": "OD",
    "Puducherry": "PY",
    "Punjab": "PB",
    "Rajasthan": "RJ",
    "Sikkim": "SK",
    "Tamil Nadu": "TN",
    "Telangana": "TS",
    "Tripura": "TR",
    "Uttar Pradesh": "UP",
    "Uttarakhand": "UK",
    "West Bengal": "WB",
}

CAPITAL_DISTRICTS = {
    "Maharashtra": "Mumbai",
    "Haryana": "Chandigarh",
    "Punjab": "Chandigarh",
    "Delhi": "New Delhi",
    "Karnataka": "Bengaluru Urban",
    "Tamil Nadu": "Chennai",
    "West Bengal": "Kolkata",
    "Gujarat": "Gandhinagar",
    "Rajasthan": "Jaipur",
    "Uttar Pradesh": "Lucknow",
    "Madhya Pradesh": "Bhopal",
    "Chandigarh": "Chandigarh",
}


def now_iso():
    return datetime.now().replace(microsecond=0).isoformat()


def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def password_hash(password):
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def public_user(row):
    return {"id": row["id"], "name": row["name"], "email": row["email"], "role": row["role"]}


def normalize(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def safe_state_code(state):
    state = normalize(state)
    if state in STATE_CODES:
        return STATE_CODES[state]
    letters = re.sub(r"[^A-Za-z]", "", state.upper())
    return (letters[:2] or "XX").ljust(2, "X")


def get_district_code(conn, state, district):
    state = normalize(state) or "Unknown"
    district = normalize(district) or "Unknown"
    row = conn.execute(
        "SELECT code FROM district_codes WHERE state=? AND district=?", (state, district)
    ).fetchone()
    if row:
        return row["code"]

    existing = conn.execute(
        "SELECT district, code FROM district_codes WHERE state=? ORDER BY code", (state,)
    ).fetchall()
    capital = CAPITAL_DISTRICTS.get(state)
    if not existing and capital:
        first = capital
        conn.execute(
            "INSERT OR IGNORE INTO district_codes(state, district, code) VALUES (?, ?, ?)",
            (state, first, "01"),
        )
        if district == first:
            return "01"
        existing = conn.execute(
            "SELECT district, code FROM district_codes WHERE state=? ORDER BY code", (state,)
        ).fetchall()

    if not existing:
        code = "01"
    else:
        code = f"{max(int(r['code']) for r in existing) + 1:02d}"
    conn.execute(
        "INSERT OR IGNORE INTO district_codes(state, district, code) VALUES (?, ?, ?)",
        (state, district, code),
    )
    return code


def next_school_code(conn, state, district):
    prefix = safe_state_code(state) + get_district_code(conn, state, district)
    row = conn.execute(
        "SELECT school_code FROM schools WHERE school_code LIKE ? ORDER BY school_code DESC LIMIT 1",
        (prefix + "%",),
    ).fetchone()
    next_num = 1
    if row and row["school_code"][-4:].isdigit():
        next_num = int(row["school_code"][-4:]) + 1
    return f"{prefix}{next_num:04d}"


def init_db():
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = db()
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS sessions (
            token TEXT PRIMARY KEY,
            user_id INTEGER NOT NULL,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS district_codes (
            state TEXT NOT NULL,
            district TEXT NOT NULL,
            code TEXT NOT NULL,
            UNIQUE(state, district)
        );
        CREATE TABLE IF NOT EXISTS schools (
            id INTEGER PRIMARY KEY,
            school_code TEXT UNIQUE,
            data_json TEXT NOT NULL,
            disposition TEXT DEFAULT 'Not Called',
            assigned_to INTEGER,
            current_campaign_id INTEGER,
            registered_event INTEGER DEFAULT 0,
            event_name TEXT DEFAULT '',
            last_note TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS edit_history (
            id INTEGER PRIMARY KEY,
            school_id INTEGER NOT NULL,
            field TEXT NOT NULL,
            old_value TEXT,
            new_value TEXT,
            user_id INTEGER,
            changed_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS reminders (
            id INTEGER PRIMARY KEY,
            school_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            due_at TEXT NOT NULL,
            note TEXT NOT NULL,
            done INTEGER DEFAULT 0,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS campaigns (
            id INTEGER PRIMARY KEY,
            name TEXT NOT NULL,
            description TEXT DEFAULT '',
            created_by INTEGER NOT NULL,
            created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS campaign_schools (
            campaign_id INTEGER NOT NULL,
            school_id INTEGER NOT NULL,
            assigned_to INTEGER,
            status TEXT DEFAULT 'Assigned',
            UNIQUE(campaign_id, school_id)
        );
        CREATE TABLE IF NOT EXISTS call_logs (
            id INTEGER PRIMARY KEY,
            school_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            campaign_id INTEGER,
            disposition TEXT NOT NULL,
            note TEXT DEFAULT '',
            created_at TEXT NOT NULL
        );
        """
    )
    users = [
        ("Admin", "admin@schoolcrm.local", "admin123", "admin"),
        ("Outreach One", "outreach1@schoolcrm.local", "outreach123", "outreach"),
        ("Outreach Two", "outreach2@schoolcrm.local", "outreach123", "outreach"),
    ]
    for name, email, password, role in users:
        conn.execute(
            "INSERT OR IGNORE INTO users(name, email, password_hash, role) VALUES (?, ?, ?, ?)",
            (name, email, password_hash(password), role),
        )
    conn.commit()
    if conn.execute("SELECT COUNT(*) c FROM schools").fetchone()["c"] == 0 and SAMPLE_XLSX.exists():
        import_workbook(conn, SAMPLE_XLSX)
    conn.close()


def school_payload(row):
    data = json.loads(row["data_json"])
    data["SCHOOL CODE"] = row["school_code"] or data.get("SCHOOL CODE", "")
    data.update(
        {
            "id": row["id"],
            "disposition": row["disposition"],
            "assigned_to": row["assigned_to"],
            "current_campaign_id": row["current_campaign_id"],
            "registered_event": bool(row["registered_event"]),
            "event_name": row["event_name"],
            "last_note": row["last_note"],
            "created_at": row["created_at"],
            "updated_at": row["updated_at"],
        }
    )
    return data


def import_workbook(conn, path_or_file):
    wb = load_workbook(path_or_file, data_only=True, read_only=True)
    ws = wb.active
    header = [normalize(v) for v in next(ws.iter_rows(min_row=1, max_row=1, max_col=FIELD_LIMIT, values_only=True))]
    field_names = [h or FIELDS[i] for i, h in enumerate(header[:FIELD_LIMIT])]
    imported = 0
    for row in ws.iter_rows(min_row=2, max_col=FIELD_LIMIT, values_only=True):
        data = {field_names[i]: normalize(row[i]) for i in range(min(len(row), FIELD_LIMIT))}
        if not any(data.values()):
            continue
        for field in FIELDS:
            data.setdefault(field, "")
        state = data.get("STATE/ UT", "")
        district = data.get("DISTRICT", "")
        code = data.get("SCHOOL CODE") or next_school_code(conn, state, district)
        data["SCHOOL CODE"] = code
        exists = conn.execute("SELECT id FROM schools WHERE school_code=?", (code,)).fetchone()
        if exists:
            continue
        conn.execute(
            """
            INSERT INTO schools(school_code, data_json, created_at, updated_at)
            VALUES (?, ?, ?, ?)
            """,
            (code, json.dumps(data), now_iso(), now_iso()),
        )
        imported += 1
    conn.commit()
    return imported


class Handler(BaseHTTPRequestHandler):
    def send_json(self, payload, status=200):
        body = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def read_json(self):
        length = int(self.headers.get("Content-Length", "0"))
        if not length:
            return {}
        return json.loads(self.rfile.read(length).decode("utf-8"))

    def current_user(self):
        token = self.headers.get("Authorization", "").replace("Bearer ", "")
        if not token:
            jar = cookies.SimpleCookie(self.headers.get("Cookie"))
            token = jar.get("session").value if jar.get("session") else ""
        if not token:
            return None
        conn = db()
        row = conn.execute(
            """
            SELECT users.* FROM sessions
            JOIN users ON users.id=sessions.user_id
            WHERE sessions.token=?
            """,
            (token,),
        ).fetchone()
        conn.close()
        return dict(row) if row else None

    def require_user(self):
        user = self.current_user()
        if not user:
            self.send_json({"error": "Login required"}, 401)
        return user

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path == "/":
            return self.serve_file(ROOT / "public" / "index.html", "text/html")
        if parsed.path == "/health":
            return self.send_json({"ok": True})
        if parsed.path.startswith("/assets/"):
            target = ROOT / "public" / parsed.path.lstrip("/")
            mime = "text/css" if target.suffix == ".css" else "application/javascript"
            return self.serve_file(target, mime)
        if parsed.path.startswith("/api/"):
            return self.handle_api_get(parsed)
        self.send_error(404)

    def serve_file(self, path, mime):
        if not path.exists():
            self.send_error(404)
            return
        data = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", mime)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def do_POST(self):
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/"):
            return self.handle_api_post(parsed)
        self.send_error(404)

    def do_PATCH(self):
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/"):
            return self.handle_api_patch(parsed)
        self.send_error(404)

    def handle_api_get(self, parsed):
        user = self.current_user() if parsed.path != "/api/me" else self.require_user()
        if parsed.path == "/api/me":
            if user:
                self.send_json({"user": public_user(user), "fields": FIELDS})
            return
        user = self.require_user()
        if not user:
            return
        conn = db()
        query = parse_qs(parsed.query)
        if parsed.path == "/api/users":
            rows = conn.execute("SELECT id, name, email, role FROM users ORDER BY role, name").fetchall()
            self.send_json({"users": [dict(r) for r in rows]})
        elif parsed.path == "/api/schools":
            clauses, params = [], []
            if user["role"] == "outreach":
                clauses.append("(assigned_to=? OR id IN (SELECT school_id FROM campaign_schools WHERE assigned_to=?))")
                params += [user["id"], user["id"]]
            if query.get("search", [""])[0]:
                clauses.append("data_json LIKE ?")
                params.append(f"%{query['search'][0]}%")
            if query.get("disposition", [""])[0]:
                clauses.append("disposition=?")
                params.append(query["disposition"][0])
            if query.get("campaign", [""])[0]:
                clauses.append("id IN (SELECT school_id FROM campaign_schools WHERE campaign_id=?)")
                params.append(query["campaign"][0])
            for key, values in query.items():
                if not key.startswith("field."):
                    continue
                field = key[6:]
                value = values[0].strip()
                if field in FIELDS and value:
                    clauses.append("data_json LIKE ?")
                    params.append(f'%"{field}": "%{value}%')
            where = " WHERE " + " AND ".join(clauses) if clauses else ""
            rows = conn.execute(f"SELECT * FROM schools{where} ORDER BY updated_at DESC LIMIT 500", params).fetchall()
            self.send_json({"schools": [school_payload(r) for r in rows]})
        elif parsed.path == "/api/filter-options":
            options = {field: set() for field in FIELDS}
            for row in conn.execute("SELECT data_json FROM schools").fetchall():
                data = json.loads(row["data_json"])
                for field in FIELDS:
                    value = normalize(data.get(field, ""))
                    if value:
                        options[field].add(value)
            self.send_json({"options": {k: sorted(v)[:250] for k, v in options.items()}})
        elif parsed.path == "/api/map":
            state_name = query.get("state", [""])[0]
            state_counts = {}
            district_counts = {}
            for row in conn.execute("SELECT data_json FROM schools").fetchall():
                data = json.loads(row["data_json"])
                state_value = normalize(data.get("STATE/ UT", "")) or "Unknown"
                district_value = normalize(data.get("DISTRICT", "")) or "Unknown"
                state_counts[state_value] = state_counts.get(state_value, 0) + 1
                if not state_name or state_value == state_name:
                    district_counts[district_value] = district_counts.get(district_value, 0) + 1
            self.send_json(
                {
                    "states": [{"state": k, "count": v} for k, v in sorted(state_counts.items())],
                    "districts": [{"district": k, "count": v} for k, v in sorted(district_counts.items())],
                    "selected_state": state_name,
                }
            )
        elif parsed.path.startswith("/api/schools/") and parsed.path.endswith("/history"):
            school_id = int(parsed.path.split("/")[3])
            field = query.get("field", [""])[0]
            params = [school_id]
            field_clause = ""
            if field:
                field_clause = " AND field=?"
                params.append(field)
            rows = conn.execute(
                f"""
                SELECT edit_history.*, users.name user_name FROM edit_history
                LEFT JOIN users ON users.id=edit_history.user_id
                WHERE school_id=?{field_clause}
                ORDER BY changed_at DESC LIMIT 100
                """,
                params,
            ).fetchall()
            self.send_json({"history": [dict(r) for r in rows]})
        elif parsed.path == "/api/campaigns":
            campaigns = []
            for row in conn.execute("SELECT * FROM campaigns ORDER BY created_at DESC").fetchall():
                total = conn.execute("SELECT COUNT(*) c FROM campaign_schools WHERE campaign_id=?", (row["id"],)).fetchone()["c"]
                done = conn.execute(
                    "SELECT COUNT(*) c FROM campaign_schools cs JOIN schools s ON s.id=cs.school_id WHERE cs.campaign_id=? AND s.disposition!='Not Called'",
                    (row["id"],),
                ).fetchone()["c"]
                campaigns.append({**dict(row), "total": total, "contacted": done})
            self.send_json({"campaigns": campaigns})
        elif parsed.path == "/api/reminders":
            clauses, params = ["done=0"], []
            if user["role"] == "outreach":
                clauses.append("user_id=?")
                params.append(user["id"])
            rows = conn.execute(
                f"""
                SELECT reminders.*, schools.school_code, schools.data_json
                FROM reminders JOIN schools ON schools.id=reminders.school_id
                WHERE {" AND ".join(clauses)}
                ORDER BY due_at ASC
                """,
                params,
            ).fetchall()
            items = []
            for r in rows:
                item = dict(r)
                item["school_name"] = json.loads(r["data_json"]).get("SCHOOL NAME", "")
                del item["data_json"]
                items.append(item)
            self.send_json({"reminders": items})
        elif parsed.path == "/api/reports":
            self.send_json(build_reports(conn))
        elif parsed.path == "/api/activity":
            if user["role"] != "admin":
                conn.close()
                return self.send_json({"error": "Admin access required"}, 403)
            self.send_json(build_activity(conn, query))
        else:
            self.send_error(404)
        conn.close()

    def handle_api_post(self, parsed):
        if parsed.path == "/api/login":
            payload = self.read_json()
            conn = db()
            row = conn.execute("SELECT * FROM users WHERE email=?", (payload.get("email", "").lower(),)).fetchone()
            if not row or row["password_hash"] != password_hash(payload.get("password", "")):
                conn.close()
                return self.send_json({"error": "Invalid email or password"}, 403)
            token = secrets.token_urlsafe(32)
            conn.execute("INSERT INTO sessions(token, user_id, created_at) VALUES (?, ?, ?)", (token, row["id"], now_iso()))
            conn.commit()
            conn.close()
            return self.send_json({"token": token, "user": public_user(row)})
        if parsed.path == "/api/signup":
            payload = self.read_json()
            name = normalize(payload.get("name"))
            email = normalize(payload.get("email")).lower()
            password = payload.get("password", "")
            role = payload.get("role", "outreach")
            if role not in {"admin", "outreach"}:
                return self.send_json({"error": "Choose admin or outreach"}, 400)
            if role == "admin" and normalize(payload.get("admin_secret")) != ADMIN_SECRET:
                return self.send_json({"error": "Admin secret code is incorrect"}, 403)
            if not name or not email or len(password) < 6:
                return self.send_json({"error": "Name, email, and a 6+ character password are required"}, 400)
            conn = db()
            try:
                cur = conn.execute(
                    "INSERT INTO users(name, email, password_hash, role) VALUES (?, ?, ?, ?)",
                    (name, email, password_hash(password), role),
                )
                token = secrets.token_urlsafe(32)
                conn.execute("INSERT INTO sessions(token, user_id, created_at) VALUES (?, ?, ?)", (token, cur.lastrowid, now_iso()))
                conn.commit()
                row = conn.execute("SELECT * FROM users WHERE id=?", (cur.lastrowid,)).fetchone()
            except sqlite3.IntegrityError:
                conn.close()
                return self.send_json({"error": "An account with this email already exists"}, 409)
            conn.close()
            return self.send_json({"token": token, "user": public_user(row)})
        user = self.require_user()
        if not user:
            return
        conn = db()
        if parsed.path == "/api/upload":
            data = self.read_multipart_file()
            count = import_workbook(conn, io.BytesIO(data))
            self.send_json({"imported": count})
        elif parsed.path == "/api/campaigns":
            payload = self.read_json()
            cur = conn.execute(
                "INSERT INTO campaigns(name, description, created_by, created_at) VALUES (?, ?, ?, ?)",
                (payload.get("name"), payload.get("description", ""), user["id"], now_iso()),
            )
            campaign_id = cur.lastrowid
            assignee = payload.get("assigned_to")
            for school_id in payload.get("school_ids", []):
                conn.execute(
                    "INSERT OR REPLACE INTO campaign_schools(campaign_id, school_id, assigned_to, status) VALUES (?, ?, ?, 'Assigned')",
                    (campaign_id, school_id, assignee),
                )
                conn.execute("UPDATE schools SET assigned_to=?, current_campaign_id=?, updated_at=? WHERE id=?", (assignee, campaign_id, now_iso(), school_id))
            conn.commit()
            self.send_json({"campaign_id": campaign_id})
        elif parsed.path == "/api/reminders":
            payload = self.read_json()
            conn.execute(
                "INSERT INTO reminders(school_id, user_id, due_at, note, created_at) VALUES (?, ?, ?, ?, ?)",
                (payload["school_id"], payload.get("user_id") or user["id"], payload["due_at"], payload.get("note", ""), now_iso()),
            )
            conn.commit()
            self.send_json({"ok": True})
        elif parsed.path == "/api/calls":
            payload = self.read_json()
            conn.execute(
                "INSERT INTO call_logs(school_id, user_id, campaign_id, disposition, note, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                (payload["school_id"], user["id"], payload.get("campaign_id"), payload["disposition"], payload.get("note", ""), now_iso()),
            )
            conn.execute(
                "UPDATE schools SET disposition=?, last_note=?, updated_at=? WHERE id=?",
                (payload["disposition"], payload.get("note", ""), now_iso(), payload["school_id"]),
            )
            conn.commit()
            self.send_json({"ok": True})
        else:
            self.send_error(404)
        conn.close()

    def handle_api_patch(self, parsed):
        user = self.require_user()
        if not user:
            return
        if not parsed.path.startswith("/api/schools/"):
            self.send_error(404)
            return
        school_id = int(parsed.path.split("/")[3])
        payload = self.read_json()
        conn = db()
        row = conn.execute("SELECT * FROM schools WHERE id=?", (school_id,)).fetchone()
        if not row:
            conn.close()
            return self.send_json({"error": "School not found"}, 404)
        data = json.loads(row["data_json"])
        field = payload.get("field")
        value = normalize(payload.get("value"))
        if field in FIELDS:
            old = data.get(field, "")
            data[field] = value
            school_code = value if field == "SCHOOL CODE" else row["school_code"]
            conn.execute("UPDATE schools SET data_json=?, school_code=?, updated_at=? WHERE id=?", (json.dumps(data), school_code, now_iso(), school_id))
        elif field in {"disposition", "assigned_to", "registered_event", "event_name", "last_note"}:
            old = normalize(row[field])
            conn.execute(
                f"UPDATE schools SET {field}=?, updated_at=? WHERE id=?",
                (value, now_iso(), school_id),
            )
        else:
            conn.close()
            return self.send_json({"error": "Field cannot be edited"}, 400)
        conn.execute(
            "INSERT INTO edit_history(school_id, field, old_value, new_value, user_id, changed_at) VALUES (?, ?, ?, ?, ?, ?)",
            (school_id, field, old, value, user["id"], now_iso()),
        )
        conn.commit()
        updated = conn.execute("SELECT * FROM schools WHERE id=?", (school_id,)).fetchone()
        conn.close()
        self.send_json({"school": school_payload(updated)})

    def read_multipart_file(self):
        content_type = self.headers.get("Content-Type", "")
        boundary = content_type.split("boundary=")[-1].encode()
        body = self.rfile.read(int(self.headers.get("Content-Length", "0")))
        parts = body.split(b"--" + boundary)
        for part in parts:
            if b"filename=" in part:
                return part.split(b"\r\n\r\n", 1)[1].rsplit(b"\r\n", 1)[0]
        return b""


def build_reports(conn):
    today = datetime.now().date()
    windows = {
        "daily": today.isoformat(),
        "weekly": (today - timedelta(days=7)).isoformat(),
        "monthly": (today - timedelta(days=30)).isoformat(),
    }
    report = {}
    for name, start in windows.items():
        calls = conn.execute("SELECT COUNT(*) c FROM call_logs WHERE date(created_at)>=date(?)", (start,)).fetchone()["c"]
        registered = conn.execute("SELECT COUNT(*) c FROM schools WHERE registered_event=1 AND date(updated_at)>=date(?)", (start,)).fetchone()["c"]
        report[name] = {"calls": calls, "registrations": registered}
    dispositions = conn.execute("SELECT disposition, COUNT(*) count FROM schools GROUP BY disposition ORDER BY count DESC").fetchall()
    campaigns = conn.execute(
        """
        SELECT c.id, c.name, COUNT(cs.school_id) total,
        SUM(CASE WHEN s.disposition!='Not Called' THEN 1 ELSE 0 END) contacted,
        SUM(CASE WHEN s.registered_event=1 THEN 1 ELSE 0 END) registered
        FROM campaigns c
        LEFT JOIN campaign_schools cs ON cs.campaign_id=c.id
        LEFT JOIN schools s ON s.id=cs.school_id
        GROUP BY c.id ORDER BY c.created_at DESC
        """
    ).fetchall()
    report["dispositions"] = [dict(r) for r in dispositions]
    report["campaigns"] = [dict(r) for r in campaigns]
    return report


def build_activity(conn, query):
    user_id = query.get("user_id", [""])[0]
    days = int(query.get("days", ["30"])[0] or 30)
    start = (datetime.now().date() - timedelta(days=days)).isoformat()
    user_filter = " AND user_id=?" if user_id else ""
    params = [start] + ([user_id] if user_id else [])

    events = []
    calls = conn.execute(
        f"""
        SELECT call_logs.*, users.name user_name, schools.school_code, schools.data_json
        FROM call_logs
        JOIN users ON users.id=call_logs.user_id
        JOIN schools ON schools.id=call_logs.school_id
        WHERE date(call_logs.created_at)>=date(?){user_filter}
        ORDER BY call_logs.created_at DESC LIMIT 200
        """,
        params,
    ).fetchall()
    for row in calls:
        data = json.loads(row["data_json"])
        events.append(
            {
                "type": "Call",
                "user_id": row["user_id"],
                "user_name": row["user_name"],
                "school_code": row["school_code"],
                "school_name": data.get("SCHOOL NAME", ""),
                "summary": f"{row['disposition']} - {row['note']}".strip(" -"),
                "created_at": row["created_at"],
            }
        )

    edits = conn.execute(
        f"""
        SELECT edit_history.*, users.name user_name, schools.school_code, schools.data_json
        FROM edit_history
        LEFT JOIN users ON users.id=edit_history.user_id
        JOIN schools ON schools.id=edit_history.school_id
        WHERE date(edit_history.changed_at)>=date(?){user_filter}
        ORDER BY edit_history.changed_at DESC LIMIT 200
        """,
        params,
    ).fetchall()
    for row in edits:
        data = json.loads(row["data_json"])
        events.append(
            {
                "type": "Edit",
                "user_id": row["user_id"],
                "user_name": row["user_name"] or "System",
                "school_code": row["school_code"],
                "school_name": data.get("SCHOOL NAME", ""),
                "summary": f"{row['field']}: {row['old_value'] or 'blank'} -> {row['new_value'] or 'blank'}",
                "created_at": row["changed_at"],
            }
        )

    reminders = conn.execute(
        f"""
        SELECT reminders.*, users.name user_name, schools.school_code, schools.data_json
        FROM reminders
        JOIN users ON users.id=reminders.user_id
        JOIN schools ON schools.id=reminders.school_id
        WHERE date(reminders.created_at)>=date(?){user_filter}
        ORDER BY reminders.created_at DESC LIMIT 200
        """,
        params,
    ).fetchall()
    for row in reminders:
        data = json.loads(row["data_json"])
        events.append(
            {
                "type": "Reminder",
                "user_id": row["user_id"],
                "user_name": row["user_name"],
                "school_code": row["school_code"],
                "school_name": data.get("SCHOOL NAME", ""),
                "summary": f"Due {row['due_at']} - {row['note']}".strip(" -"),
                "created_at": row["created_at"],
            }
        )

    events.sort(key=lambda event: event["created_at"], reverse=True)
    by_user = {}
    for event in events:
        bucket = by_user.setdefault(event["user_name"], {"calls": 0, "edits": 0, "reminders": 0})
        if event["type"] == "Call":
            bucket["calls"] += 1
        elif event["type"] == "Edit":
            bucket["edits"] += 1
        elif event["type"] == "Reminder":
            bucket["reminders"] += 1
    return {"events": events[:300], "summary": by_user}


if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", "8766"))
    host = os.environ.get("HOST", "127.0.0.1")
    print(f"School CRM running at http://{host}:{port}")
    ThreadingHTTPServer((host, port), Handler).serve_forever()
